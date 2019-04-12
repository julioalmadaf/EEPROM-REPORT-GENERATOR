##################
#### Librerias ###
##################
import os
import sys
import tkinter as tk
import xml.etree.ElementTree as ET
import xml.etree.ElementTree as e

from os import listdir
from tkinter import messagebox
from tkinter import filedialog
from tkinter import ttk
from tkinter import *

try:
    import xlrd
except ImportError:
    os.system('python -m pip install xlrd')

try:
    import xlwt
except ImportError:
    os.system('python -m pip install xlwt')

try:
    from pathlib import Path
except ImportError:
    os.system('python -m pip install pathlib')

try:
    import shutil
except ImportError:
    os.system('python -m pip install pytest-shutil')

try:
    from PIL import ImageTk, Image
except ImportError:
    os.system('python -m pip install pillow')

try:
    from openpyxl import load_workbook
except ImportError:
    os.system('python -m pip install openpyxl')

try:
    import pandas as pd
except ImportError:
    os.system('python -m pip install pandas')

try:
    import win32com.client
except ImportError:
    os.system('python -m pip install pypiwin32')
    os.system('python -m pip install pywin32')

###########################
#### Variables globales ###
###########################
rutaCNT = "0"
rutaArchivoCNT = "0"
archivoCNTCargado = 0
rutaReportePrevio = "0"
archivoReportePrevioCargado = 0
BBNumber = 0
Baseline = 0
estadoCheckButton = 0
estadoEtiqueta = 1

################
#### Eventos ###
################
def newProgram():

    #Configuracion inicial
    estadoAgregarArchivoCNT()

def exitProgram():
        
    global estadoEtiqueta

    label.configure(text="Closing program")

    #Preguntar al usuario si desea salir del programa.
    salir = messagebox.askyesno(message="Do you want to close the program?", title="Close program")

    if(salir == 1):
            sys.exit(0)
    
    if(estadoEtiqueta == 1):
    	label.configure(text="EEPROM report generator")
    elif(estadoEtiqueta == 2):
    	label.configure(text="CNT file loaded")
    elif(estadoEtiqueta == 3):
    	label.configure(text="CNT file loaded - Previous Report loaded")
    else:	#Si entra aqui es porque algo grave paso
    	label.configure(text="EEPROM report generator")

def aboutProgram():
        
    global estadoEtiqueta

    label.configure(text="About the program")

    messagebox.showinfo("About EEPROM report generator", "This software has been released by Julio Cesar Almada Fuerte and Ruben Barajas Curiel")

    if(estadoEtiqueta == 1):
    	label.configure(text="EEPROM report generator")
    elif(estadoEtiqueta == 2):
    	label.configure(text="CNT file loaded")
    elif(estadoEtiqueta == 3):
    	label.configure(text="CNT file loaded - Previous Report loaded")
    else:	#Si entra aqui es porque algo grave paso
    	label.configure(text="EEPROM report generator")

def helpProgram():
        
    global estadoEtiqueta

    label.configure(text="Help")

    messagebox.showinfo("Help", "Visit the following link to get more information about this software")

    if(estadoEtiqueta == 1):
    	label.configure(text="EEPROM report generator")
    elif(estadoEtiqueta == 2):
    	label.configure(text="CNT file loaded")
    elif(estadoEtiqueta == 3):
    	label.configure(text="CNT file loaded - Previous Report loaded")
    else:	#Si entra aqui es porque algo grave paso
    	label.configure(text="EEPROM report generator")

def cntButton():

    global rutaCNT
    global archivoCNTCargado
    global estadoEtiqueta
    global rutaReportePrevio
    global archivoReportePrevioCargado
    global estadoCheckButton

    #Abrir Dialog Box para buscar el archivo.
    label.configure(text="Opening CNT file")
    rutaCNT = filedialog.askopenfilename(filetypes = (("All CNT files","*.cnt"),("All files","*.*")))

    #Verificar que sea archivo CNT o que se haya agregado un archivo.
    #No se selecciono un archivo CNT
    if(rutaCNT.find(".cnt") == -1):
        messagebox.showerror("Error", "Not .cnt file selected")
        #Comenzar programa desde cero.
        estadoAgregarArchivoCNT()

    #Se dio al boton cancelar.
    elif(rutaCNT == ""):
        messagebox.showerror("Error", "Not .cnt file selected")
		#Comenzar programa desde cero.
        estadoAgregarArchivoCNT()

    #El archivo seleccionado no corresponde al nombre de formato de CNT
    elif(rutaCNT.find("EEPROM_Container") == -1):
    	messagebox.showerror("Error", "Not CNT name format file")
    	#Comenzar programa desde cero.
    	estadoAgregarArchivoCNT()

    #Se selecciono el archivo correctamente.
    else:
        archivoCNTCargado = 1
        estadoGenerarReporteSinPrevio()

def enableButtonRP():

    global estadoCheckButton
    global archivoReportePrevioCargado

    estadoCheckButton = estadoCheckButton ^ 1

    #Cambiar estado del boton de reporte previo dependiendo del CheckButton.
    if(estadoCheckButton == 0):
            button_PreviousReport.state(["disabled"])
    else:
            button_PreviousReport.state(["!disabled"])

    #Si ya existe un reporte cargado, esconder o mostrar el label actualizado con la accion a hacer.
    if(archivoReportePrevioCargado == 1 and estadoCheckButton == 1):
    	label.configure(text="CNT file loaded - Previous Report loaded")
    	estadoEtiqueta = 3	#La etiqueta muestra el tercer mensaje (CNT file loaded - Previous Report loaded)
    if(archivoReportePrevioCargado == 1 and estadoCheckButton == 0):
    	label.configure(text="CNT file loaded")
    	estadoEtiqueta = 2	#La etiqueta muestra el segundo mensaje (CNT file loaded)

def previousReport():

    global rutaReportePrevio
    global archivoReportePrevioCargado
    global estadoEtiqueta
    global rutaReportePrevio

    #Abrir Dialog Box para buscar el archivo.
    label.configure(text="Opening previous report file")
    rutaReportePrevio = filedialog.askopenfilename(filetypes = (("All Excel files","*.xlsx"),("All files","*.*")))

    #Verificar que sea archivo XLSX o que se haya agregado un archivo.
    if(rutaReportePrevio.find(".xlsx") == -1):       #No se selecciono un archivo XLXS
        messagebox.showerror("Error", "Not .xlsx file selected")
        archivoReportePrevioCargado = 0
        label.configure(text="CNT file loaded")
        estadoEtiqueta = 2	#La etiqueta muestra el segundo mensaje (CNT file loaded)

    #Se dio al boton cancelar.
    elif(rutaReportePrevio== ""):
        messagebox.showerror("Error", "Not .xlsx file selected")
        archivoReportePrevioCargado = 0
        label.configure(text="CNT file loaded")
        estadoEtiqueta = 2	#La etiqueta muestra el segundo mensaje (CNT file loaded)

    #El archivo seleccionado no corresponde al nombre de formato de reporte previo
    elif(rutaReportePrevio.find("EEPROM_Container_Review_Checkist") == -1):
        messagebox.showerror("Error", "Not previous report name format file")
        archivoReportePrevioCargado = 0
        label.configure(text="CNT file loaded")
        estadoEtiqueta = 2	#La etiqueta muestra el segundo mensaje (CNT file loaded)

    #Se selecciono el archivo correctamente.
    else:
    	#Para evitar archivos excel con un nombre valido, verificar que las celdas internas correspondan a un archivo del formato deseado.
    	file_plantilla = xlrd.open_workbook(filename = "EEPROM_Container_Review_Template.xlsx")
    	sheet_plantilla = file_plantilla.sheet_by_index(0)
    	file_intento = xlrd.open_workbook(filename = rutaReportePrevio)
    	sheet_intento = file_intento.sheet_by_index(0)

    	file_revision = 0
    	for column in range (0, 17):
    		if(sheet_plantilla.cell_value(10, column) != sheet_intento.cell_value(10, column)):
    			file_revision = 1
    			break
    	if(file_revision == 0):
    		archivoReportePrevioCargado = 1
    		label.configure(text="CNT file loaded - Previous Report loaded")
    		estadoEtiqueta = 3	#La etiqueta muestra el tercer mensaje (CNT file loaded - Previous Report loaded)
    	else:
    		messagebox.showerror("Error", "This file is not a previous report")
    		archivoReportePrevioCargado = 0
    		label.configure(text="CNT file loaded")
    		estadoEtiqueta = 2	#La etiqueta muestra el segundo mensaje (CNT file loaded)

def createReport():

    global archivoCNTCargado

    label.configure(text="Creating report")

    #Garantizar que se haya seleccionado un archivo CNT.
    if(archivoCNTCargado == 1):                 
        archivoCNTCargado = 0

        #Rellena Excel
        fillExcel()

        messagebox.showinfo("Report created", "Report created successfully")

    else:
        messagebox.showerror("Error", "Not .cnt file selected")

    #Como ya se termino la actividad, arrancar con una nueva.
    estadoAgregarArchivoCNT()

################
#### Objetos ###
################
root = Tk()

rutaActual = os.getcwd()
img = ImageTk.PhotoImage(Image.open(rutaActual + "/bosch.png"))

panelElements = ttk.Frame(root, padding=(3,3,12,12))
panelImage = ttk.Frame(panelElements, borderwidth=5, relief="sunken", width=200, height=200)

label = ttk.Label(panelElements, text="EEPROM report generator", font=("Tahoma", 25, 'bold'))
button_CNT = ttk.Button(panelElements, text="Select CNT file", style="TButton", command=cntButton)
button_PreviousReport = ttk.Button(panelElements, text="Select previous report", style="TButton", command=previousReport)
button_GenerateReport = ttk.Button(panelElements, text="Generate report", style="TButton", command=createReport)
enable_button = Checkbutton(panelElements, text="Enable previous report button", onvalue=1,offvalue=0, command=enableButtonRP)
image = tk.Label(panelImage, image = img)
menubar = Menu(panelImage)

button_style1 = ttk.Style()
button_style2 = ttk.Style()

##################
#### Funciones ###
##################
def fillExcel():

    global rutaArchivoCNT
    global estadoCheckButton
    global rutaCNT
    global rutaArchivoCNT
    global BBNumber
    global Baseline

    #Extraer BBNumber y Baseline.
    BBNumber = rutaCNT.split('_')[3]
    Baseline = rutaCNT.split('_')[4]

    #Crear archivo Excel.
    rutaArchivoCNT = os.path.dirname(rutaCNT)
    shutil.copy("EEPROM_Container_Review_Template.xlsx", rutaArchivoCNT + "/fillexcel.xlsx")
    
    #Open log file.
    logProgram = open(rutaArchivoCNT + "/logProgram" + str(BBNumber) + ".txt","w+")

    logProgram.write("Generating new Excel file\r\n\r\n")

    #Carga el archivo Excel anteriormente generado.
    wb = load_workbook(filename = rutaArchivoCNT + "/fillexcel.xlsx")
    ws = wb.active
    
    #Lee archivo XML.
    tree = ET.parse(rutaCNT)
    logProgram.write("CNT file read.\r\n\r\n")

    #Obtiene el root del XML.
    root = tree.getroot()
    
    #Counter para ir agregando elementos en excel.
    CounterFilasExcel = 11
    
    ######################################################
    #Guarda el nombre del proyecto.
    for project in root.iter('PROJECT-INFO'):
        logProgram.write("PROJECT-INFO\r\n")
        PD = project.find('PROJECT-DESC')
        if PD is not None:
                logProgram.write("      PROJECT-DESC -- " + PD.text + "\r\n")

    ######################################################
    #Guarda el nombre del responsable.
    for info in root.iter('RESPONSIBLE'):
        logProgram.write("RESPONSIBLE \r\n")
        PN = info.find('PERSON-NAME')
        if PN is not None:
                logProgram.write("      PERSON-NAME -- " + PN.text + "\r\n")

    ######################################################
    #Busca el nodo sesion en todo el arbol.
    logProgram.write("SESSIONS\r\n")
    for session in root.iter('SESSION'):

        #Busca en los tipos de sesiones que nombre tiene.
        sessionN = session.find('SESSION-NAME')
        if sessionN is not None:
                logProgram.write("      SESSION-NAME -- " + sessionN.text + "\r\n")

                #Cuando la sesion es ALL.
                if(sessionN.text == '__ALL__'):

                        #Para no alterar el orden de las filas del excel.
                        tempCounter=CounterFilasExcel
                        
                        #Obtiene el Datapointer-name del item.
                        for DPN in session.iter('DATAPOINTER-NAME'):
                                logProgram.write("                      DATAPOINTER-NAME -- " + str(DPN.text) + "\r\n")
                                tempCounter += 1
                                ws['A'+ str(tempCounter)] = DPN.text    #Guarda el valor en el excel.
                        
                        tempCounter = CounterFilasExcel
                        
                        #Obtiene el Datapointer-ident del item.
                        for DPID in session.iter('DATAPOINTER-IDENT'):
                                logProgram.write("                      DATAPOINTER-ID -- " + str(DPID.text) + "\r\n")
                                tempCounter += 1
                                ws['B'+ str(tempCounter)] = DPID.text   #Guarda el valor en el excel.
                        
                        tempCounter = CounterFilasExcel
                        
                        #Obtiene el Datapointer-identifier del item.
                        for DFID in session.iter('DATAFORMAT-IDENTIFIER'):
                                logProgram.write("                      DATAPOINTER-DESIRED TYPE -- " + str(DFID.text) + "\r\n")
                                CounterFilasExcel += 1          #Aqui se aumenta el CounterFilasExcel para que se respeten las filas.
                                ws['M'+ str(CounterFilasExcel)] = DFID.text

                #Cuando la sesion es Reprog.
                if(sessionN.text == 'Reprog'):

                        tempCounter = CounterFilasExcel
                        
                        for DPN in session.iter('DATAPOINTER-NAME'):
                                logProgram.write("                      DATAPOINTER-NAME -- " + str(DPN.text) + "\r\n")
                                tempCounter += 1
                                ws['A'+ str(tempCounter)] = DPN.text
                        
                        tempCounter = CounterFilasExcel
                        
                        for DPID in session.iter('DATAPOINTER-IDENT'):
                                logProgram.write("                      DATAPOINTER-ID -- " + str(DPID.text) + "\r\n")
                                tempCounter += 1
                                ws['B'+ str(tempCounter)] = DPID.text
                        
                        tempCounter=CounterFilasExcel
                        
                        for DFID in session.iter('DATAFORMAT-IDENTIFIER'):
                                logProgram.write("                      DATAPOINTER-DESIRED TYPE -- " + str(DFID.text) + "\r\n")
                                CounterFilasExcel += 1
                                ws['M'+ str(CounterFilasExcel)] = DFID.text
                                ws['I'+ str(CounterFilasExcel)] = "X"           #Marca que el use case de que es Reprog.

                #Cuando la sesion es DeliveryState.
                if(sessionN.text == 'DeliveryState'):

                        tempCounter = CounterFilasExcel

                        for DPN in session.iter('DATAPOINTER-NAME'):
                                logProgram.write("                      DATAPOINTER-NAME -- " + str(DPN.text) + "\r\n")
                                tempCounter += 1
                                ws['A'+ str(tempCounter)] = DPN.text

                        tempCounter = CounterFilasExcel

                        for DPID in session.iter('DATAPOINTER-IDENT'):
                                logProgram.write("                      DATAPOINTER-ID -- " + str(DPID.text) + "\r\n")
                                tempCounter += 1
                                ws['B'+ str(tempCounter)] = DPID.text
                        
                        tempCounter = CounterFilasExcel

                        for DFID in session.iter('DATAFORMAT-IDENTIFIER'):
                                logProgram.write("                      DATAPOINTER-DESIRED TYPE -- " + str(DFID.text) + "\r\n")
                                CounterFilasExcel += 1
                                ws['M'+ str(CounterFilasExcel)] = DFID.text
                                ws['G'+ str(CounterFilasExcel)] = "X"           #Marca que el use case de que es DeliveryState.

                #Cuando es la sesion es ResetToDeliveryState.
                if(sessionN.text == 'ResetToDeliveryState'):

                        tempCounter = CounterFilasExcel

                        for DPN in session.iter('DATAPOINTER-NAME'):
                                logProgram.write("                      DATAPOINTER-NAME -- " + str(DPN.text) + "\r\n")
                                tempCounter += 1
                                ws['A'+ str(tempCounter)] = DPN.text
                        
                        tempCounter = CounterFilasExcel

                        for DPID in session.iter('DATAPOINTER-IDENT'):
                                logProgram.write("                      DATAPOINTER-ID -- " + str(DPID.text) + "\r\n")
                                tempCounter += 1
                                ws['B'+ str(tempCounter)] = DPID.text
                        
                        tempCounter = CounterFilasExcel

                        for DFID in session.iter('DATAFORMAT-IDENTIFIER'):
                                logProgram.write("                      DATAPOINTER-DESIRED TYPE -- " + str(DFID.text) + "\r\n")
                                CounterFilasExcel += 1
                                ws['M'+ str(CounterFilasExcel)] = DFID.text
                                ws['H'+ str(CounterFilasExcel)] = "X"           #Marca que el use case de que es ReturnToDeliveryState.

    logProgram.write("Repeated Datapointers deleted " + str(DPID.text) + "\r\n")
    
    ######################################################
    #Para checar si se repite algun NVM Item.
    for i in range(12,CounterFilasExcel):
        #Agarra cada fila y las compara con las demas.
        temp = ws['A' + str(i)]
        k = i + 1
        for j in range(k, CounterFilasExcel):
            #Aqui se toma el siguiente en la fila y se checa cada elemento siguiente.
            temp2 = ws['A' + str(j)]
            #Si son iguales.
            if(temp.value == temp2.value):
                #Checa el USE CASES de cada uno.
                temp3 = ws['G' + str(j)]
                if(temp3.value == "X"):
                    #Marca el use case del que se repite .
                    ws['G'+str(i)]="X"
                    #Borra la fila que se repite.
                    ws.delete_rows(j,1)
                    CounterFilasExcel-=1
                temp3 = ws['H'+str(j)]
                if(temp3.value=="X"): 
                    ws['H'+str(i)]="X"
                    ws.delete_rows(j,1)
                    CounterFilasExcel-=1
                temp3 = ws['I'+str(j)]
                if(temp3.value=="X"): 
                    ws['I'+str(i)]="X"
                    ws.delete_rows(j,1)
                    CounterFilasExcel-=1

    ######################################################
    #Agrega comments.
    for datablock in root.iter('DATABLOCK'):
        #Busca en los datablock que nombre tiene.
        DBN = datablock.find('DATABLOCK-NAME')
        #Para recorrer el excel.
        for i in range(12, CounterFilasExcel):
            #Compara el valor que tiene la celda de excel con el datablock name.
            if(DBN.text == ((ws['A' + str(i)].value) + '__Metadata')):
                j = 0
                #Hay varios DATA por datablock, pero el que se ocupa es el 6to.
                for DPN in datablock.iter('DATA'):
                    j += 1
                    if(j == 6):
                        logProgram.write("DATAPOINTER-NAME -- " + DBN.text + "\r\n")
                        logProgram.write("       COMMENT -- " + str(DPN.text) + "\r\n")
                        #Se copia la descripcion a la columna comment.
                        ws['O' + str(i)] = DPN.text


    #Si un reporte previo es agregado y se tiene habilitada la opcion.
    if(archivoReportePrevioCargado == 1 and estadoCheckButton == 1):
        logFile = open(rutaArchivoCNT + "/logFile" + str(BBNumber) + ".txt","w+")
        sheet1=wb.worksheets[0]
        #Cuenta las filas maximas que tiene el archivo original.
        newCounterFilasExcel=sheet1.max_row
        wb2=load_workbook(rutaReportePrevio)
        sheet2=wb2.worksheets[0]
        ws2=wb2.active
        #Cuenta cuantos elementos tiene el archivo previo seleccionado.
        row_count = sheet2.max_row
        for i in range(12, newCounterFilasExcel):
            for j in range(12, row_count):
                if(ws['A'+str(i)].value==ws2['A'+str(j)].value):
                    logFile.write("      FROM " + ws['A'+str(i)].value + " \r\n")
                    #ID Number.
                    if(ws['B'+str(i)].value==ws2['B'+str(j)].value):
                        ws['B'+str(i)]=ws2['B'+str(j)].value
                        logFile.write("          ID NUMBER NOT CHANGED \r\n")
                    else:
                        ws['B'+str(i)]=ws2['B'+str(j)].value
                        logFile.write("          ID NUMBER CHANGED \r\n")

                    #cr-p.
                    if(ws['C'+str(i)].value==ws2['C'+str(j)].value):
                        ws['C'+str(i)]=ws2['C'+str(j)].value
                        logFile.write("          CR - P NOT CHANGED \r\n")
                    else:
                        ws['C'+str(i)]=ws2['C'+str(j)].value
                        logFile.write("          CR - P CHANGED \r\n")
                    
                    #CRP delivery state.
                    if(ws['D'+str(i)].value==ws2['D'+str(j)].value):
                        ws['D'+str(i)]=ws2['D'+str(j)].value
                        logFile.write("          CRP DELIVERY STATE NOT CHANGED \r\n")
                    else:
                        ws['D'+str(i)]=ws2['D'+str(j)].value
                        logFile.write("          CRP DELIVERY STATE CHANGED \r\n")

                    #CRP reset delivery state.
                    if(ws['E'+str(i)].value==ws2['E'+str(j)].value):
                        ws['E'+str(i)]=ws2['E'+str(j)].value
                        logFile.write("          CRP RESET DELIVERY STATE NOT CHANGED \r\n")
                    else:
                        ws['E'+str(i)]=ws2['E'+str(j)].value
                        logFile.write("          CRP RESET DELIVERY STATE CHANGED \r\n")
                    
                    #CRP reprog.
                    if(ws['F'+str(i)].value==ws2['F'+str(j)].value):
                        ws['F'+str(i)]=ws2['F'+str(j)].value
                        logFile.write("          CRP REPROG NOT CHANGED \r\n")
                    else:
                        ws['F'+str(i)]=ws2['F'+str(j)].value
                        logFile.write("          CRP REPROG CHANGED \r\n")
                    
                    #Expected delivery state.
                    if(ws['J'+str(i)].value==ws2['J'+str(j)].value):
                        ws['J'+str(i)]=ws2['J'+str(j)].value
                        logFile.write("          EXPECTED DELIVERY STATE NOT CHANGED \r\n")
                    else:
                        ws['J'+str(i)]=ws2['J'+str(j)].value
                        logFile.write("          EXPECTED DELIVERY STATE CHANGED \r\n")
                    
                    #Expected reset delivery state.
                    if(ws['K'+str(i)].value==ws2['K'+str(j)].value):
                        ws['K'+str(i)]=ws2['K'+str(j)].value
                        logFile.write("          EXPECTED RESET DELIVERY STATE NOT CHANGED \r\n")
                    else:
                        ws['K'+str(i)]=ws2['K'+str(j)].value
                        logFile.write("          EXPECTED RESET DELIVERY STATE CHANGED \r\n")
                    
                    #Expected reprog.
                    if(ws['L'+str(i)].value==ws2['L'+str(j)].value):
                        ws['L'+str(i)]=ws2['L'+str(j)].value
                        logFile.write("          EXPECTED REPROG NOT CHANGED \r\n")
                    else:
                        ws['L'+str(i)]=ws2['L'+str(j)].value
                        logFile.write("          EXPECTED REPROG CHANGED \r\n")

                    #Desired type.
                    if(ws['M'+str(i)].value==ws2['M'+str(j)].value):
                        ws['M'+str(i)]=ws2['M'+str(j)].value
                        logFile.write("          DESIRED TYPE NOT CHANGED \r\n")
                    else:
                        ws['M'+str(i)]=ws2['M'+str(j)].value
                        logFile.write("          DESIRED TYPE CHANGED \r\n")
                    
                    #Desired data.
                    if(ws['N'+str(i)].value==ws2['N'+str(j)].value):
                        ws['N'+str(i)]=ws2['N'+str(j)].value
                        logFile.write("          DESIRED DATA NOT CHANGED \r\n")
                    else:
                        ws['N'+str(i)]=ws2['N'+str(j)].value
                        logFile.write("          DESIRED DATA CHANGED \r\n")
                    
                    #Comment.
                    if(ws['O'+str(i)].value==ws2['O'+str(j)].value):
                        ws['O'+str(i)]=ws2['O'+str(j)].value
                        logFile.write("          COMMENT NOT CHANGED \r\n")
                    else:
                        ws['O'+str(i)]=ws2['O'+str(j)].value
                        logFile.write("          COMMENT CHANGED \r\n")
                    
                    #Rating.
                    if(ws['P'+str(i)].value==ws2['P'+str(j)].value):
                        ws['P'+str(i)]=ws2['P'+str(j)].value
                        logFile.write("          RATING NOT CHANGED \r\n")
                    else:
                        ws['P'+str(i)]=ws2['P'+str(j)].value
                        logFile.write("          RATING CHANGED \r\n")
                    
                    #Rated by.
                    if(ws['Q'+str(i)].value==ws2['Q'+str(j)].value):
                        ws['Q'+str(i)]=ws2['Q'+str(j)].value
                        logFile.write("          RATED BY NOT CHANGED \r\n")
                    else:
                        ws['Q'+str(i)]=ws2['Q'+str(j)].value
                        logFile.write("          RATED BY CHANGED \r\n")
                    
                    #Comments.
                    if(ws['R'+str(i)].value==ws2['R'+str(j)].value):
                        ws['R'+str(i)]=ws2['R'+str(j)].value
                        logFile.write("          COMMENTS NOT CHANGED \r\n")
                    else:
                        ws['R'+str(i)]=ws2['R'+str(j)].value
                        logFile.write("          COMMENTS CHANGED \r\n")
                    
                    #Reference comments from GA.
                    if(ws['S'+str(i)].value==ws2['S'+str(j)].value):
                        ws['S'+str(i)]=ws2['S'+str(j)].value
                        logFile.write("          REFERENCE COMMENTS FROM GA NOT CHANGED \r\n")
                    else:
                        ws['S'+str(i)]=ws2['S'+str(j)].value
                        logFile.write("          REFERENCE COMMENTS FROM GA CHANGED \r\n")
        logFile.close()
    
    #Guarda los cambios.
    wb.save(rutaArchivoCNT + "/fillexcel.xlsx")

    #Para ordenar por ID number. Selecciona el archivo excel.
    logProgram.write("\r\nEXCEL FILE DATA POINTERS SORTED BY NVM ID NUMBER\r\n\r\n")
    excel_file = rutaArchivoCNT + "/fillexcel.xlsx"
    #Leer el archivo.
    movies = pd.read_excel(excel_file, skiprows=10)
    #Ordena por numero de ID.
    sorted_by_number = movies.sort_values(by='ID number',ascending=True)
    #Guardar.
    sorted_by_number.to_excel(excel_file,index=False)

    #Se crea una copia del Template para poder copiar los datos ordenados al archivo que se generara al final.
    shutil.copy("EEPROM_Container_Review_Template.xlsx", rutaArchivoCNT + "/EEPROM_Container_Review_Checkist_GM_iPB_GlobalB_" + BBNumber + ".xlsx")
    
    #Carga el archivo Excel anteriormente generado.
    wb1 = load_workbook(filename = rutaArchivoCNT +  "/EEPROM_Container_Review_Checkist_GM_iPB_GlobalB_" + BBNumber + ".xlsx")
    ws1=wb1.active

    #Carga el archivo con los datos ordenados.
    wb=load_workbook(filename = rutaArchivoCNT +  "/fillexcel.xlsx")
    sheet=wb.worksheets[0]
    
    #Para que no escriba en espacios vacios.
    row_count = sheet.max_row
    ws=wb.active

    #Desde aqui toma los datos.
    j=2

    #Los datos los pega ordenados en el archivo excel que es copia del template.
    #No escribe en espacios vacios.
    for i in range(12, 12+row_count-1):
        ws1['A'+str(i)]=ws['A'+str(j)].value
        ws1['B'+str(i)]=ws['B'+str(j)].value
        ws1['C'+str(i)]=ws['C'+str(j)].value
        ws1['D'+str(i)]=ws['D'+str(j)].value
        ws1['E'+str(i)]=ws['E'+str(j)].value
        ws1['F'+str(i)]=ws['F'+str(j)].value
        ws1['G'+str(i)]=ws['G'+str(j)].value
        ws1['H'+str(i)]=ws['H'+str(j)].value
        ws1['I'+str(i)]=ws['I'+str(j)].value
        ws1['J'+str(i)]=ws['J'+str(j)].value
        ws1['K'+str(i)]=ws['K'+str(j)].value
        ws1['L'+str(i)]=ws['L'+str(j)].value
        ws1['M'+str(i)]=ws['M'+str(j)].value
        ws1['N'+str(i)]=ws['N'+str(j)].value
        ws1['O'+str(i)]=ws['O'+str(j)].value
        ws1['P'+str(i)]=ws['P'+str(j)].value
        ws1['Q'+str(i)]=ws['Q'+str(j)].value
        ws1['R'+str(i)]=ws['R'+str(j)].value
        ws1['S'+str(i)]=ws['S'+str(j)].value
        j+=1

    #Asigna los valores de BBNumber, Baseline, Encargado y nombre del proyecto a sus respectivas celdas.
    logProgram.write("BBNumber -- " + str(BBNumber) + " added to Excel file.\r\n\r\n")
    logProgram.write("Baseline -- " + str(Baseline) + " added to Excel file.\r\n\r\n")
    logProgram.write("PROJECT MANAGER -- " + PN.text + " added to Excel file.\r\n\r\n")
    logProgram.write("PROJECT DESCRIPTION -- " + PD.text + " added to Excel file.\r\n\r\n")
    ws1['B3']=BBNumber
    ws1['B4']=Baseline
    ws1['B5']=PN.text
    ws1['B2']=PD.text

    #Guarda el archivo.
    wb1.save(rutaArchivoCNT +  "/EEPROM_Container_Review_Checkist_GM_iPB_GlobalB_" + BBNumber + ".xlsx")
    
    #Borra el archivo que tiene los datos ordenados.
    os.remove(rutaArchivoCNT + "/fillexcel.xlsx")

    #Close log.
    logProgram.write("END OF LOG FILE\r\n")
    logProgram.close()

def estadoAgregarArchivoCNT():
	
    global rutaCNT
    global rutaArchivoCNT
    global archivoCNTCargado
    global rutaReportePrevio
    global archivoReportePrevioCargado
    global estadoCheckButton
    global estadoEtiqueta

    #Ocultar botones innecesarios (reporte previo, checkButton y generar reporte)
    button_PreviousReport.grid_remove()
    enable_button.grid_remove()
    button_GenerateReport.grid_remove()

    #Dar enfoque al boton de agregar archivo CNT.
    button_CNT.configure(style='button_style1.TButton')
    button_PreviousReport.configure(style='button_style2.TButton')
    button_GenerateReport.configure(style='button_style2.TButton')

    #Comenzar con el boton de reporte previo deshabilitado y deseleccionado.
    estadoCheckButton = 0
    button_PreviousReport.state(["disabled"])
    enable_button.deselect()

    #Mostrar etiqueta de paso inicial.
    label.configure(text="EEPROM report generator")
    estadoEtiqueta = 1	#La etiqueta muestra el primer mensaje (EEPROM report generator)

    rutaCNT = "0"
    rutaArchivoCNT = "0"
    archivoCNTCargado = 0
    rutaReportePrevio = "0"
    archivoReportePrevioCargado = 0
    BBNumber = 0
    Baseline = 0
    estadoCheckButton = 0
    estadoEtiqueta = 1

def estadoGenerarReporteSinPrevio():
	
    global estadoEtiqueta

    #Acomodar fuente de los botones. Resaltar boton de generar reporte.
    button_CNT.configure(style='button_style2.TButton')
    button_GenerateReport.configure(style='button_style1.TButton')

    #Mostrar botones necesarios para la siguiente actividad (generar reporte, habilitar reporte previo y buscar reporte previo)
    enable_button.grid()            #Ahora se puede mostrar el checkbuton para habilitar el boton de reporte previo.
    button_PreviousReport.state(["disabled"])
    button_PreviousReport.grid()    #Ahora se puede mostrar el boton de reporte previo como opcional.
    button_GenerateReport.grid()    #Ahora se puede mostrar el boton de generar reporte.

    #Mostrar etiqueta de segundo paso.
    label.configure(text="CNT file loaded")
    estadoEtiqueta = 2	#La etiqueta muestra el segundo mensaje (CNT file loaded)

def main():
                
    #Titulo de la ventana.
    root.title("Bosch") 

    #Configurar paneles.
    panelElements.grid(column=0, row=0, sticky=(N, S, E, W))
    panelImage.grid(column=0, row=1, columnspan=2, rowspan=6, sticky=(N, S, E, W))

    #Configurar elementos (botones, etiqueta, imagen, etc).
    image.pack(side = "bottom", fill = "both", expand = "yes")
    label.grid(column=0, row=0, columnspan=4, sticky=(N, W))
    button_CNT.grid(column=3, row=3)
    enable_button.grid(column=3, row=4)
    button_PreviousReport.grid(column=3, row=5)
    button_GenerateReport.grid(column=3, row=6)
        
    #Ocultar botones innecesarios.
    button_PreviousReport.grid_remove()
    enable_button.grid_remove()
    button_GenerateReport.grid_remove()

    #Fondos y colores.
    style = ttk.Style(root)
    style.configure('TLabel', background='white')       #Background y foreground de la etiqueta.
    style.configure('TFrame', background='white')       #Background y foreground del Frame.

    #Estilo de los elementos.
    button_style1.configure("button_style1.TButton", width = 20, padding=5, font=('Helvetica', 10, 'bold'), background = "black", foreground = 'green')
    button_style2.configure("button_style2.TButton", width = 20, padding=5, font=('Helvetica', 10, 'bold'))
        
    button_CNT.configure(style='button_style1.TButton')
    button_PreviousReport.configure(style='button_style2.TButton')
    button_GenerateReport.configure(style='button_style2.TButton')

    #Comenzar con el boton de reporte previo deshabilitado.
    estadoCheckButton = 0
    button_PreviousReport.state(["disabled"])

    root.config(menu=menubar)
    toolBar_Init = Menu(menubar)
    toolBar_About = Menu(menubar)
    toolBar_Init.add_command(label="New", command=newProgram)
    toolBar_Init.add_command(label="Exit", command=exitProgram)
    toolBar_About.add_command(label="About", command=aboutProgram)
    toolBar_About.add_command(label="Help", command=helpProgram)
    menubar.add_cascade(label="File", menu=toolBar_Init)
    menubar.add_cascade(label="Program", menu=toolBar_About)

    #Comenzar proceso.
    root.mainloop()

#################################
if __name__== "__main__":
    main()
